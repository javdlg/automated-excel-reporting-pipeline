import pandas as pd
import logging
import os
from typing import Dict
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment

# Setup basic logging to console
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def load_data(data_dict: Dict[str, pd.DataFrame], output_path: str) -> bool:
    """
    Generate the final formatted Excel artifact from processed DataFrames.
    
    Args:
        data_dict (Dict[str, pd.DataFrame]): Dictionary containing the DataFrames to write.
        output_path (str): The destination path for the final Excel report.
        
    Returns:
        bool: True if the loading process is successful, False otherwise.
    """
    logging.info(f"Starting to load data into Excel report at: {output_path}")
    
    try:
        # Ensure output directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Initialize a new workbook
        wb = Workbook()
        
        # Remove the default empty sheet created by openpyxl
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
            
        # Iterate over each dataframe and create a sheet
        for sheet_name, df in data_dict.items():
            ws = wb.create_sheet(title=sheet_name)
            
            # Write dataframe to worksheet (including headers, ignoring index)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
                
            # Apply visual formatting
            _format_worksheet(ws, sheet_name)
            
        # Save the finalized workbook
        wb.save(output_path)
        logging.info("Excel report successfully generated and formatted.")
        return True
        
    except Exception as e:
        logging.error(f"An error occurred while generating the Excel report: {e}")
        return False

from openpyxl.worksheet.worksheet import Worksheet

def _format_worksheet(ws: Worksheet, sheet_name: str):
    """
    Apply visual formatting to the worksheet based on the sheet type.
    
    Args:
        ws: openpyxl worksheet object.
        sheet_name (str): Name of the worksheet.
    """
    # 1. Format headers (Bold, Background Color, Alignment)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        
    # 2. Auto-adjust column widths based on content
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width
        
    # 3. Apply specific formatting based on sheet content
    if sheet_name == "Revenue_by_Branch":
        # Format total_revenue column as currency
        # Assuming total_revenue is the 3rd column (Branch, City, Total Revenue)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=3)
            # Apply currency format
            cell.number_format = '"$"#,##0.00'
            
    elif sheet_name == "Avg_Rating_by_Product":
        # Format average_rating column
        # Assuming average_rating is the 2nd column (Product Line, Average Rating)
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")
        
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=2)
            cell.number_format = '0.00'
            
            # Conditional formatting: Highlight ratings below 6.0 in red
            try:
                if float(cell.value) < 6.0:
                    cell.fill = red_fill
                    cell.font = red_font
            except (ValueError, TypeError):
                pass
